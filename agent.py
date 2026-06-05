#!/usr/bin/env python3
"""Rocasis outbound sales agent.

Local CLI for selecting qualified prospects, generating personalized outreach,
sending through Resend, and tracking confirmed Cal.com meetings.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import html
import http.server
import json
import os
import re
import sys
import textwrap
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable


ROOT = Path(__file__).resolve().parent
DEFAULT_HTML = Path("/Users/macmini/Downloads/prospectos_rocasis_nemaris_outreach_05052026_v3_corregido.html")
DEFAULT_XLSX = Path("/Users/macmini/Desktop/Prospectos/Rocasis-Nemaris-DB.xlsx")
DEFAULT_TRACKER = ROOT / "outreach_state" / "rocasis_outreach_tracker.csv"
DEFAULT_OUTBOX = ROOT / "outbox"
CAL_LINK = "https://cal.com/rocasis"
CAL_API_VERSION = "2026-05-01"
HTTP_USER_AGENT = "rocasis-sales-agent/1.0 (+https://www.rocasis.com/)"
DEFAULT_REPLY_TO = "marketing.voxmedia@gmail.com"
TARGET_MEETINGS = 4
ACTIVE_STATUSES = {"sent", "followup_1", "followup_2"}
FINAL_STATUSES = {"confirmed", "do_not_contact", "closed"}
EXCLUDED_FROM_NEW_SEND = ACTIVE_STATUSES | FINAL_STATUSES
RESEND_UNCONTACTABLE_STATUSES = {"bounced", "suppressed", "complained", "failed", "rejected"}

ROLE_TERMS = (
    "cio",
    "cto",
    "head of it",
    "head of software",
    "head of software engineering",
    "chief information officer",
    "it director",
    "it manager",
    "information security",
    "ciso",
)

TECH_ROLE_TERMS = (
    "ti",
    "it",
    "tecnologia",
    "tecnologías",
    "sistemas",
    "infraestructura",
    "transformacion digital",
    "transformación digital",
    "transformation",
    "innovacion",
    "innovación",
    "arquitectura",
    "operaciones ti",
    "gestion ti",
    "gestión ti",
    "seguridad",
    "information security",
)

SECTOR_COMPANY_TERMS = {
    "financiero": (
        "afirme",
        "ahorra seguros",
        "banco",
        "bancoppel",
        "bankaya",
        "stori",
        "capital seguros",
        "certus",
        "credito",
        "crédito",
        "financiera",
        "inbursa",
        "kapital",
        "seguros",
    ),
    "retail": (
        "arabela",
        "benotto",
        "chedraui",
        "flexi",
        "juguetron",
        "juguetrón",
        "mobo",
        "o´reilly",
        "oreilly",
        "pharmacy",
        "super",
        "farmacia",
        "gilsa",
        "liverpool",
    ),
    "manufactura": (
        "azucarero",
        "bosch",
        "cleber",
        "cuprum",
        "jumex",
        "kasto",
        "laboratorio",
        "laboratorios",
        "manufactura",
        "automotriz",
        "pharma",
    ),
}

MEXICO_HINTS = (
    ".mx",
    "mexico",
    "méxico",
    "bajio",
    "bajío",
    "bancoppel",
    "chedraui",
    "jumex",
    "juguetron",
    "juguetrón",
    "benotto",
    "cuprum",
    "afirme",
    "inbursa",
    "covalto",
    "mobo",
)

SECTOR_TERMS = {
    "financiero": ("financ", "banca", "banco", "bank", "seguros", "fintech", "credit", "card"),
    "retail": (
        "retail",
        "supermercado",
        "tienda",
        "tiendas",
        "sucursales",
        "e-commerce",
        "ecommerce",
        "accesorios",
        "autopartes",
        "acabados",
        "bicicletas",
    ),
    "manufactura": ("manufactura", "automotriz", "produccion", "planta", "dispositivos medicos", "pharma"),
}

SCORE_ORDER = {"Alta": 0, "Media-Alta": 1, "Media": 2}


class ApiRequestError(RuntimeError):
    def __init__(self, status_code: int, details: str):
        super().__init__(f"HTTP {status_code}: {details}")
        self.status_code = status_code
        self.details = details


def load_dotenv(path: Path = ROOT / ".env") -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


@dataclass
class Prospect:
    name: str
    company: str
    role: str
    email: str
    industry: str = ""
    project: str = "Rocasis"
    score: str = "Media"
    context: str = ""
    systems: str = ""
    trigger: str = ""
    angle: str = ""
    subject: str = ""
    email_body: str = ""
    linkedin: str = ""
    phone: str = ""
    sectors: list[str] = field(default_factory=list)

    @property
    def key(self) -> str:
        return self.email.strip().lower()


def now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")


def compact_spaces(value: str | None) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_key(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", ascii_value.lower()).strip()


def role_rank(role: str) -> int:
    role_l = role.lower()
    if any(term in role_l for term in ("cio", "cto", "ciso", "chief information")):
        return 0
    if "director" in role_l:
        return 1
    if "head" in role_l:
        return 2
    if "gerente" in role_l or "manager" in role_l:
        return 3
    return 9


def load_html_dataset(path: Path) -> list[Prospect]:
    raw = path.read_text(encoding="utf-8")
    match = re.search(r"const DATA = (\[.*?\]);\s*function esc", raw, re.S)
    if not match:
        raise ValueError(f"No DATA array found in {path}")
    records = json.loads(match.group(1))
    prospects: list[Prospect] = []
    for row in records:
        prospects.append(
            Prospect(
                name=compact_spaces(row.get("name")),
                company=compact_spaces(row.get("company")),
                role=compact_spaces(row.get("role")),
                email=compact_spaces(row.get("email")).lower(),
                industry=compact_spaces(row.get("industry")),
                project=compact_spaces(row.get("project")) or "Rocasis",
                score=compact_spaces(row.get("score")) or "Media",
                context=compact_spaces(row.get("context")),
                systems=compact_spaces(row.get("systems")),
                trigger=compact_spaces(row.get("trigger")),
                angle=compact_spaces(row.get("angle")),
                subject=compact_spaces(row.get("subject")),
                email_body=str(row.get("email_body") or "").strip(),
            )
        )
    return prospects


def load_xlsx_dataset(path: Path) -> list[Prospect]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover - environment fallback
        raise RuntimeError("openpyxl is required to read XLSX fallback data") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    prospects: list[Prospect] = []
    # Hoja2/Hoja3 are curated shortlists; Hoja1 is the broader database.
    sheet_names = ["Hoja2", "Hoja3", "Hoja1"]
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [compact_spaces(c.value).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            email_addr = compact_spaces(values.get("email")).lower()
            if "@" not in email_addr:
                continue
            prospects.append(
                Prospect(
                    name=compact_spaces(values.get("prospecto")),
                    company=compact_spaces(values.get("lugar de trabajo")),
                    role=compact_spaces(values.get("cargo")),
                    email=email_addr,
                    linkedin=compact_spaces(values.get("linkedin")),
                    phone=compact_spaces(values.get("celular")),
                )
            )
    return prospects


def load_prospects(html_path: Path, xlsx_path: Path) -> list[Prospect]:
    prospects: list[Prospect] = []
    if html_path.exists():
        prospects.extend(load_html_dataset(html_path))
    if xlsx_path.exists():
        prospects.extend(load_xlsx_dataset(xlsx_path))
    if prospects:
        return prospects
    raise FileNotFoundError(f"No prospect source found: {html_path} or {xlsx_path}")


def infer_sectors(prospect: Prospect) -> list[str]:
    haystack = " ".join(
        [
            prospect.industry,
            prospect.context,
            prospect.systems,
            prospect.trigger,
            prospect.angle,
            prospect.company,
            prospect.email,
        ]
    ).lower()
    sectors = []
    for sector, terms in SECTOR_TERMS.items():
        company_terms = SECTOR_COMPANY_TERMS.get(sector, ())
        if any(term_matches(haystack, term) for term in terms + company_terms):
            sectors.append(sector)
    return sectors


def term_matches(haystack: str, term: str) -> bool:
    # Short tokens need word boundaries to avoid false positives inside words.
    if len(term) <= 4 and term.isalpha():
        return re.search(rf"\b{re.escape(term)}\b", haystack) is not None
    return term in haystack


def is_decision_maker(prospect: Prospect) -> bool:
    role = prospect.role.lower()
    if "project manager" in role and not any(
        role_term_matches(role, term)
        for term in ("infraestructura", "infrastructure", "operaciones ti", "servicios ti", "information security")
    ):
        return False
    if any(role_term_matches(role, term) for term in ROLE_TERMS):
        return True
    if role_term_matches(role, "director") or role_term_matches(role, "gerente") or role_term_matches(role, "manager"):
        return any(
            role_term_matches(role, term) for term in TECH_ROLE_TERMS
        )
    return False


def role_term_matches(role: str, term: str) -> bool:
    if term in {"cio", "cto", "ciso", "ti", "it"}:
        return re.search(rf"\b{re.escape(term)}\b", role) is not None
    if len(term) <= 4 and term.isalpha():
        return re.search(rf"\b{re.escape(term)}\b", role) is not None
    return term in role


def is_mexico_prospect(prospect: Prospect) -> bool:
    haystack = f"{prospect.company} {prospect.email} {prospect.context} {prospect.industry} {prospect.role}".lower()
    return any(hint in haystack for hint in MEXICO_HINTS)


def qualify(prospects: Iterable[Prospect]) -> list[Prospect]:
    qualified: list[Prospect] = []
    seen: set[str] = set()
    for prospect in prospects:
        if not prospect.email or "@" not in prospect.email:
            continue
        if prospect.key in seen:
            continue
        if prospect.project and prospect.project.lower() != "rocasis":
            continue
        if not is_decision_maker(prospect):
            continue
        if not is_mexico_prospect(prospect):
            continue
        prospect.sectors = infer_sectors(prospect)
        if not prospect.sectors:
            continue
        seen.add(prospect.key)
        qualified.append(prospect)
    qualified.sort(key=lambda p: (SCORE_ORDER.get(p.score, 99), role_rank(p.role), p.company.lower(), p.name.lower()))
    deduped: list[Prospect] = []
    seen_identity: set[str] = set()
    for prospect in qualified:
        identity = f"{normalize_key(prospect.company)}|{normalize_key(prospect.name)}"
        if identity in seen_identity:
            continue
        seen_identity.add(identity)
        deduped.append(prospect)
    return deduped


def tracker_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def write_tracker(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "email",
        "prospect",
        "company",
        "sector",
        "status",
        "sent_at",
        "followup_1_at",
        "followup_2_at",
        "closed_at",
        "resend_id",
        "resend_status",
        "resend_last_checked_at",
        "booking_uid",
        "meeting_start",
        "meeting_confirmed_at",
        "notes",
    ]
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def tracker_index(path: Path) -> dict[str, dict[str, str]]:
    return {row.get("email", "").lower(): row for row in tracker_rows(path)}


def generate_subject(prospect: Prospect) -> str:
    if prospect.subject:
        return prospect.subject
    return f"{prospect.company}: integracion de sistemas sin reprocesos"


def generated_body(prospect: Prospect) -> str:
    systems = prospect.systems or "ERP, CRM, e-commerce, inventario, facturacion y otras aplicaciones criticas"
    context = prospect.context or "operacion, datos, procesos comerciales y sistemas criticos"
    return textwrap.dedent(
        f"""\
        Hola {first_name(prospect.name)},

        Te escribo porque en empresas como {prospect.company}, donde conviven {context}, uno de los retos mas comunes para TI es lograr que los sistemas criticos trabajen conectados sin depender de procesos manuales, integraciones fragiles o desarrollos aislados.

        Desde Roca Sistemas trabajamos con Magic xpi para ayudar a equipos de TI a orquestar e integrar {systems}, con automatizacion de procesos y conectores empresariales sin redisenar toda la arquitectura actual.

        Me gustaria proponerte una sesion breve de assessment de 30 minutos para revisar si hay algun punto de friccion donde podamos aportar valor.

        Puedes tomar el horario que mejor te funcione aqui:
        {CAL_LINK}

        Saludos,
        Miguel Cedillo

        Si no eres la persona correcta o prefieres no recibir mas mensajes, respondeme "baja" y lo retiro de seguimiento.
        """
    ).strip()


def first_name(name: str) -> str:
    return (name.strip().split() or [""])[0]


def ensure_cal_cta(body: str) -> str:
    if CAL_LINK in body:
        updated = body
    else:
        updated = body.rstrip() + "\n\nPuedes tomar el horario que mejor te funcione aqui:\n" + CAL_LINK
    if "baja" not in updated.lower():
        updated += '\n\nSi no eres la persona correcta o prefieres no recibir mas mensajes, respondeme "baja" y lo retiro de seguimiento.'
    return updated


def generate_body(prospect: Prospect) -> str:
    source_body = prospect.email_body.strip()
    if source_body:
        return ensure_cal_cta(source_body)
    return generated_body(prospect)


def text_to_html(body: str) -> str:
    paragraphs = [f"<p>{html.escape(part).replace(chr(10), '<br>')}</p>" for part in body.split("\n\n")]
    return "\n".join(paragraphs)


def selected_prospects(args: argparse.Namespace) -> list[Prospect]:
    prospects = qualify(load_prospects(Path(args.html_source), Path(args.xlsx_source)))
    tracker = tracker_index(Path(args.tracker))
    remaining = [
        p
        for p in prospects
        if tracker.get(p.key, {}).get("status") not in EXCLUDED_FROM_NEW_SEND
    ]
    return remaining[: args.limit]


def write_previews(prospects: list[Prospect], outbox: Path) -> None:
    outbox.mkdir(parents=True, exist_ok=True)
    with (outbox / "selected_prospects.csv").open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=["name", "company", "role", "email", "industry", "sectors", "score", "subject"],
        )
        writer.writeheader()
        for p in prospects:
            writer.writerow(
                {
                    "name": p.name,
                    "company": p.company,
                    "role": p.role,
                    "email": p.email,
                    "industry": p.industry,
                    "sectors": ", ".join(p.sectors),
                    "score": p.score,
                    "subject": generate_subject(p),
                }
            )
    with (outbox / "email_previews.txt").open("w", encoding="utf-8") as fh:
        for p in prospects:
            fh.write("=" * 88 + "\n")
            fh.write(f"To: {p.name} <{p.email}>\n")
            fh.write(f"Company: {p.company} | Role: {p.role} | Sector: {', '.join(p.sectors)}\n")
            fh.write(f"Subject: {generate_subject(p)}\n\n")
            fh.write(generate_body(p))
            fh.write("\n\n")


def resend_email(api_key: str, sender: str, reply_to: str, prospect: Prospect) -> str:
    return resend_message(
        api_key=api_key,
        sender=sender,
        reply_to=reply_to,
        to_email=prospect.email,
        subject=generate_subject(prospect),
        body=generate_body(prospect),
        campaign="rocasis-calcom-4-meetings",
        idempotency_key=idempotency_key("initial", prospect.email, generate_subject(prospect)),
    )


def resend_message(
    api_key: str,
    sender: str,
    reply_to: str,
    to_email: str,
    subject: str,
    body: str,
    campaign: str,
    idempotency_key: str = "",
) -> str:
    payload = {
        "from": sender,
        "to": [to_email],
        "reply_to": reply_to,
        "subject": subject,
        "text": body,
        "html": text_to_html(body),
        "headers": {"X-Campaign": campaign},
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "User-Agent": HTTP_USER_AGENT,
    }
    if idempotency_key:
        headers["Idempotency-Key"] = idempotency_key
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            data = json.loads(response.read().decode("utf-8"))
            return str(data.get("id", ""))
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Resend HTTP {exc.code}: {details}") from exc


def validate_resend_api_key(api_key: str) -> bool:
    return api_key.startswith("re_") and len(api_key) >= 12


def idempotency_key(*parts: str) -> str:
    source = "|".join(compact_spaces(part).lower() for part in parts)
    return "rocasis-" + hashlib.sha256(source.encode("utf-8")).hexdigest()[:48]


def request_json(url: str, headers: dict[str, str]) -> dict:
    headers = {"User-Agent": HTTP_USER_AGENT, **headers}
    req = urllib.request.Request(url, headers=headers, method="GET")
    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise ApiRequestError(exc.code, details) from exc


def post_json(url: str, payload: dict, headers: dict[str, str] | None = None) -> dict:
    request_headers = {"User-Agent": HTTP_USER_AGENT, "Content-Type": "application/json"}
    if headers:
        request_headers.update(headers)
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers=request_headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            raw = response.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise ApiRequestError(exc.code, details) from exc


def fetch_cal_bookings(api_key: str, statuses: list[str], api_version: str = CAL_API_VERSION) -> list[dict]:
    headers = {"Authorization": f"Bearer {api_key}", "cal-api-version": api_version}
    bookings: list[dict] = []
    for status in statuses:
        cursor = ""
        while True:
            query = f"?status={urllib.parse.quote(status)}" if status else ""
            if cursor:
                query += ("&" if query else "?") + f"cursor={urllib.parse.quote(cursor)}"
            payload = request_json(f"https://api.cal.com/v2/bookings{query}", headers)
            bookings.extend(payload.get("data") or [])
            pagination = payload.get("pagination") or {}
            if not pagination.get("hasMore") or not pagination.get("nextCursor"):
                break
            cursor = str(pagination["nextCursor"])
    return bookings


def fetch_resend_domains(api_key: str) -> list[dict]:
    payload = request_json("https://api.resend.com/domains", {"Authorization": f"Bearer {api_key}"})
    return list(payload.get("data") or [])


def fetch_resend_email(api_key: str, email_id: str) -> dict:
    return request_json(
        f"https://api.resend.com/emails/{urllib.parse.quote(email_id)}",
        {"Authorization": f"Bearer {api_key}"},
    )


def sender_domain(sender: str) -> str:
    if not sender:
        return ""
    match = re.search(r"@([^>\s]+)", sender)
    return match.group(1).lower() if match else sender.split("@")[-1].lower()


def validate_resend_live(api_key: str, sender: str = "") -> dict:
    result = {
        "ok": False,
        "domains": [],
        "enabled_domains": [],
        "sender_domain": sender_domain(sender),
        "error": "",
    }
    if not api_key:
        result["error"] = "RESEND_API_KEY is missing."
        return result
    if not validate_resend_api_key(api_key):
        result["error"] = "RESEND_API_KEY does not look like a Resend key; expected prefix re_."
        return result
    try:
        domains = fetch_resend_domains(api_key)
    except ApiRequestError as exc:
        result["error"] = f"Resend API rejected the request: HTTP {exc.status_code}: {compact_spaces(exc.details)}"
        return result
    result["domains"] = domains
    enabled = [
        str(domain.get("name", "")).lower()
        for domain in domains
        if str((domain.get("capabilities") or {}).get("sending", "")) == "enabled"
    ]
    result["enabled_domains"] = enabled
    if not domains:
        result["error"] = "No Resend domains were returned for this API key."
        return result
    if result["sender_domain"] and result["sender_domain"] not in enabled:
        result["error"] = f"Sender domain is not enabled in Resend: {result['sender_domain']}"
        return result
    result["ok"] = True
    return result


def attendee_emails(booking: dict) -> set[str]:
    emails: set[str] = set()
    for attendee in booking.get("attendees") or []:
        email_addr = compact_spaces(attendee.get("email")).lower()
        if email_addr:
            emails.add(email_addr)
    for guest in booking.get("guests") or []:
        email_addr = compact_spaces(guest).lower()
        if email_addr:
            emails.add(email_addr)
    return emails


def attendee_names(booking: dict) -> list[str]:
    names: list[str] = []
    for attendee in booking.get("attendees") or []:
        name = compact_spaces(attendee.get("name"))
        if name:
            names.append(name)
    return names


def extract_booking(payload: dict) -> dict:
    data = payload.get("data") if isinstance(payload.get("data"), dict) else payload
    booking = data.get("booking") if isinstance(data.get("booking"), dict) else data
    return booking if isinstance(booking, dict) else {}


def known_prospect_index(args: argparse.Namespace) -> dict[str, Prospect]:
    return {p.key: p for p in qualify(load_prospects(Path(args.html_source), Path(args.xlsx_source)))}


def mark_booking_confirmed(
    rows: list[dict[str, str]],
    prospects: dict[str, Prospect],
    booking: dict,
    note: str,
) -> int:
    index = {row.get("email", "").lower(): row for row in rows}
    matched = 0
    for email_addr in attendee_emails(booking):
        prospect = prospects.get(email_addr)
        if not prospect and email_addr not in index:
            continue
        row = index.get(email_addr)
        if not row:
            row = {"email": email_addr}
            rows.append(row)
            index[email_addr] = row
        row.update(
            {
                "email": email_addr,
                "prospect": row.get("prospect") or (prospect.name if prospect else ""),
                "company": row.get("company") or (prospect.company if prospect else ""),
                "sector": row.get("sector") or (", ".join(prospect.sectors) if prospect else ""),
                "status": "confirmed",
                "booking_uid": compact_spaces(booking.get("uid")),
                "meeting_start": compact_spaces(booking.get("start")),
                "meeting_confirmed_at": now_iso(),
                "notes": row.get("notes") or note,
            }
        )
        matched += 1
    return matched


def unmatched_booking_rows(bookings: list[dict], matched_booking_uids: set[str]) -> list[dict[str, str]]:
    rows = []
    for booking in bookings:
        uid = compact_spaces(booking.get("uid"))
        if uid in matched_booking_uids:
            continue
        rows.append(
            {
                "booking_uid": uid,
                "start": compact_spaces(booking.get("start")),
                "status": compact_spaces(booking.get("status")),
                "attendee_emails": ", ".join(sorted(attendee_emails(booking))),
                "attendee_names": ", ".join(attendee_names(booking)),
                "title": compact_spaces(booking.get("title")),
            }
        )
    return rows


def write_cal_unmatched_report(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = ["booking_uid", "start", "status", "attendee_emails", "attendee_names", "title"]
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def next_followup_step(row: dict[str, str], today: dt.date | None = None) -> str:
    today = today or dt.datetime.now(dt.timezone.utc).date()
    status = row.get("status", "")
    sent_at = parse_date(row.get("sent_at"))
    followup_1_at = parse_date(row.get("followup_1_at"))
    if status == "sent" and sent_at and (today - sent_at).days >= 2:
        return "followup_1"
    if status == "followup_1" and followup_1_at and (today - followup_1_at).days >= 3:
        return "followup_2"
    if status == "followup_2" and parse_date(row.get("followup_2_at")) and (today - parse_date(row.get("followup_2_at"))).days >= 5:
        return "closed"
    return ""


def parse_date(value: str | None) -> dt.date | None:
    if not value:
        return None
    try:
        return dt.datetime.fromisoformat(value.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def followup_subject(row: dict[str, str], step: str) -> str:
    company = row.get("company") or "tu empresa"
    if step == "followup_1":
        return f"Re: {company}: integracion de sistemas"
    if step == "followup_2":
        return f"Cierro el ciclo por ahora, {company}"
    return f"{company}: seguimiento"


def followup_body(row: dict[str, str], step: str) -> str:
    name = first_name(row.get("prospect", ""))
    company = row.get("company") or "tu empresa"
    if step == "followup_1":
        return textwrap.dedent(
            f"""\
            Hola {name},

            Retomo mi mensaje anterior. La razon de contactarte es revisar si en {company} hay fricciones entre ERP, CRM, e-commerce, inventario, facturacion u otros sistemas criticos que esten generando reprocesos o poca visibilidad operativa.

            En Roca Sistemas usamos Magic xpi para integrar aplicaciones empresariales y automatizar flujos sin redisenar toda la arquitectura actual.

            Si te hace sentido, puedes agendar una sesion de 30 minutos aqui:
            {CAL_LINK}

            Si no lo ves contigo, ¿quien lleva integraciones, automatizacion o arquitectura de aplicaciones?

            Saludos,
            Miguel Cedillo
            """
        ).strip()
    return textwrap.dedent(
        f"""\
        Hola {name},

        Cierro el ciclo por ahora para no insistir de mas. Si en {company} estan revisando integracion de sistemas, automatizacion de procesos o conectividad entre ERP/CRM/e-commerce/SAP, aqui dejo el enlace para una sesion breve:

        {CAL_LINK}

        Si no eres la persona correcta o prefieres no recibir mas mensajes, respondeme "baja" y lo retiro de seguimiento.

        Saludos,
        Miguel Cedillo
        """
    ).strip()


def due_followups(rows: list[dict[str, str]], today: dt.date | None = None) -> list[tuple[dict[str, str], str]]:
    due: list[tuple[dict[str, str], str]] = []
    for row in rows:
        step = next_followup_step(row, today=today)
        if step:
            due.append((row, step))
    return due


def cmd_prepare(args: argparse.Namespace) -> int:
    prospects = selected_prospects(args)
    write_previews(prospects, Path(args.outbox))
    print(f"Prepared {len(prospects)} qualified Rocasis prospects.")
    print(f"Previews: {Path(args.outbox) / 'email_previews.txt'}")
    print(f"Shortlist: {Path(args.outbox) / 'selected_prospects.csv'}")
    return 0


def cmd_send(args: argparse.Namespace) -> int:
    prospects = selected_prospects(args)
    if not prospects:
        print("No unsent qualified prospects remain.")
        return 0
    write_previews(prospects, Path(args.outbox))
    if not args.send:
        print(f"Dry-run: prepared {len(prospects)} emails. Add --send to send through Resend.")
        return 0
    api_key = os.environ.get("RESEND_API_KEY", "").strip()
    if not api_key:
        print("RESEND_API_KEY is missing from the environment.", file=sys.stderr)
        return 2
    if not validate_resend_api_key(api_key):
        print("RESEND_API_KEY does not look like a Resend API key. Expected prefix: re_", file=sys.stderr)
        return 2
    if not args.sender or not args.reply_to:
        print("--from and --reply-to are required when sending.", file=sys.stderr)
        return 2

    rows = tracker_rows(Path(args.tracker))
    index = {row.get("email", "").lower(): row for row in rows}
    sent = 0
    for prospect in prospects:
        resend_id = resend_email(api_key, args.sender, args.reply_to, prospect)
        row = index.get(prospect.key)
        if not row:
            row = {"email": prospect.email}
            rows.append(row)
            index[prospect.key] = row
        row.update(
            {
                "email": prospect.email,
                "prospect": prospect.name,
                "company": prospect.company,
                "sector": ", ".join(prospect.sectors),
                "status": "sent",
                "sent_at": now_iso(),
                "resend_id": resend_id,
                "notes": row.get("notes", ""),
            }
        )
        sent += 1
        print(f"Sent {sent}/{len(prospects)}: {prospect.company} - {prospect.name}")
        if args.delay_seconds:
            time.sleep(args.delay_seconds)
    write_tracker(Path(args.tracker), rows)
    return 0


def cmd_followups(args: argparse.Namespace) -> int:
    rows = tracker_rows(Path(args.tracker))
    due = due_followups(rows)
    if args.limit:
        due = due[: args.limit]
    outbox = Path(args.outbox)
    outbox.mkdir(parents=True, exist_ok=True)
    preview_path = outbox / "followup_previews.txt"
    with preview_path.open("w", encoding="utf-8") as fh:
        for row, step in due:
            fh.write("=" * 88 + "\n")
            fh.write(f"To: {row.get('prospect')} <{row.get('email')}>\n")
            fh.write(f"Company: {row.get('company')} | Step: {step}\n")
            fh.write(f"Subject: {followup_subject(row, step)}\n\n")
            fh.write(followup_body(row, step))
            fh.write("\n\n")
    if not due:
        print("No follow-ups are due.")
        return 0
    if not args.send:
        print(f"Dry-run: prepared {len(due)} follow-ups. Preview: {preview_path}")
        return 0
    api_key = os.environ.get("RESEND_API_KEY", "").strip()
    if not api_key:
        print("RESEND_API_KEY is missing from the environment.", file=sys.stderr)
        return 2
    if not validate_resend_api_key(api_key):
        print("RESEND_API_KEY does not look like a Resend API key. Expected prefix: re_", file=sys.stderr)
        return 2
    if not args.sender or not args.reply_to:
        print("--from and --reply-to are required when sending.", file=sys.stderr)
        return 2
    for row, step in due:
        if step == "closed":
            row["status"] = "closed"
            row["closed_at"] = now_iso()
            print(f"Closed sequence: {row.get('company')} - {row.get('prospect')}")
            continue
        resend_id = resend_message(
            api_key=api_key,
            sender=args.sender,
            reply_to=args.reply_to,
            to_email=row["email"],
            subject=followup_subject(row, step),
            body=followup_body(row, step),
            campaign=f"rocasis-calcom-{step}",
            idempotency_key=idempotency_key(step, row["email"], followup_subject(row, step)),
        )
        row["status"] = step
        row[f"{step}_at"] = now_iso()
        row["resend_id"] = resend_id
        print(f"Sent {step}: {row.get('company')} - {row.get('prospect')}")
        if args.delay_seconds:
            time.sleep(args.delay_seconds)
    write_tracker(Path(args.tracker), rows)
    return 0


def cmd_check_resend(args: argparse.Namespace) -> int:
    api_key = os.environ.get("RESEND_API_KEY", "").strip()
    result = validate_resend_live(api_key, args.sender)
    for domain in result["domains"]:
        name = str(domain.get("name", "")).lower()
        status = str(domain.get("status", ""))
        sending = str((domain.get("capabilities") or {}).get("sending", ""))
        print(f"{name}: status={status}, sending={sending}")
    if not result["ok"]:
        print(result["error"], file=sys.stderr)
        return 1 if result["domains"] else 2
    return 0


def extract_resend_status(payload: dict) -> str:
    for key in ("last_event", "status", "state"):
        value = payload.get(key)
        if value:
            return compact_spaces(value).lower()
    events = payload.get("events")
    if isinstance(events, list) and events:
        latest = events[-1]
        if isinstance(latest, dict):
            return compact_spaces(latest.get("type") or latest.get("event") or latest.get("name")).lower()
        return compact_spaces(latest).lower()
    return "unknown"


def apply_resend_status(row: dict[str, str], resend_status: str) -> None:
    row["resend_status"] = resend_status
    row["resend_last_checked_at"] = now_iso()
    if resend_status in RESEND_UNCONTACTABLE_STATUSES:
        row["status"] = "do_not_contact"
        note = f"Marked do_not_contact after Resend status: {resend_status}"
        row["notes"] = "; ".join(part for part in [row.get("notes", ""), note] if part)


def cmd_sync_resend(args: argparse.Namespace) -> int:
    api_key = os.environ.get("RESEND_API_KEY", "").strip()
    if not api_key:
        print("RESEND_API_KEY is missing from the environment.", file=sys.stderr)
        return 2
    if not validate_resend_api_key(api_key):
        print("RESEND_API_KEY does not look like a Resend API key. Expected prefix: re_", file=sys.stderr)
        return 2
    rows = tracker_rows(Path(args.tracker))
    checked = 0
    failed = 0
    for row in rows:
        resend_id = row.get("resend_id", "")
        if not resend_id:
            continue
        try:
            payload = fetch_resend_email(api_key, resend_id)
        except ApiRequestError as exc:
            failed += 1
            row["resend_status"] = f"error_http_{exc.status_code}"
            row["resend_last_checked_at"] = now_iso()
            if not args.quiet:
                print(f"Failed {row.get('email')}: HTTP {exc.status_code}")
            continue
        apply_resend_status(row, extract_resend_status(payload))
        checked += 1
        if not args.quiet:
            print(f"{row.get('email')}: {row['resend_status']}")
        if args.delay_seconds:
            time.sleep(args.delay_seconds)
    write_tracker(Path(args.tracker), rows)
    print(f"Synced Resend status for {checked} emails. Failed: {failed}")
    return 0 if failed == 0 else 1


def cmd_sync_cal(args: argparse.Namespace) -> int:
    api_key = os.environ.get("CAL_API_KEY", "").strip()
    if not api_key:
        print("CAL_API_KEY is missing from the environment.", file=sys.stderr)
        return 2
    result = sync_cal_bookings(
        api_key=api_key,
        tracker=Path(args.tracker),
        outbox=Path(args.outbox),
        prospects=known_prospect_index(args),
        statuses=args.status,
        api_version=args.api_version,
    )
    print(result["message"])
    print(f"Unmatched report: {result['report_path']}")
    if result["matched"] and args.notify:
        send_notification(f"Cal.com sync: {result['matched']} booking(s) matched for Rocasis.")
    return 0


def sync_cal_bookings(
    api_key: str,
    tracker: Path,
    outbox: Path,
    prospects: dict[str, Prospect],
    statuses: list[str],
    api_version: str = CAL_API_VERSION,
) -> dict:
    rows = tracker_rows(tracker)
    bookings = fetch_cal_bookings(api_key, statuses, api_version=api_version)
    matched = 0
    matched_booking_uids: set[str] = set()
    for booking in bookings:
        count = mark_booking_confirmed(rows, prospects, booking, "Synced from Cal.com bookings API")
        matched += count
        if count:
            matched_booking_uids.add(compact_spaces(booking.get("uid")))
    write_tracker(tracker, rows)
    unmatched = unmatched_booking_rows(bookings, matched_booking_uids)
    report_path = outbox / "cal_unmatched_bookings.csv"
    write_cal_unmatched_report(report_path, unmatched)
    return {
        "fetched": len(bookings),
        "matched": matched,
        "unmatched": len(unmatched),
        "report_path": str(report_path),
        "message": f"Fetched {len(bookings)} Cal.com bookings. Matched: {matched}. Unmatched: {len(unmatched)}.",
    }


def send_telegram_message(token: str, chat_id: str, text: str) -> dict:
    return post_json(
        f"https://api.telegram.org/bot{urllib.parse.quote(token)}/sendMessage",
        {"chat_id": chat_id, "text": text, "disable_web_page_preview": True},
    )


def fetch_telegram_updates(token: str, offset: int = 0, timeout: int = 25) -> list[dict]:
    payload = {"timeout": timeout, "allowed_updates": ["message"]}
    if offset:
        payload["offset"] = offset
    response = post_json(f"https://api.telegram.org/bot{urllib.parse.quote(token)}/getUpdates", payload)
    return list(response.get("result") or [])


def send_notification(text: str) -> bool:
    token = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
    chat_id = os.environ.get("TELEGRAM_CHAT_ID", "").strip()
    if not token or not chat_id:
        print("Telegram notification skipped: TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID is missing.")
        return False
    send_telegram_message(token, chat_id, text)
    return True


def telegram_help_text() -> str:
    return "\n".join(
        [
            "Rocasis agent commands:",
            "/status - campaign progress",
            "/sync_cal - sync Cal.com bookings",
            "/doctor - readiness and next action",
            "/help - show commands",
        ]
    )


def build_doctor_text(args: argparse.Namespace) -> str:
    report = build_doctor_report(args)
    checks = report["checks"]
    lines = [
        "Rocasis Sales Agent Doctor",
        f"- Qualified prospects: {checks['qualified_prospects']}",
        f"- Confirmed meetings: {checks['confirmed_meetings']}/{checks['target_meetings']}",
        f"- Sent awaiting booking: {checks['sent']}",
        f"- Follow-ups due: {checks['due_followups']}",
        f"- Resend key present: {checks['resend_key_present']}",
        f"- Resend key format ok: {checks['resend_key_format_ok']}",
        f"- Cal.com key present: {checks['cal_key_present']}",
    ]
    if report["blockers"]:
        lines.append("Blockers:")
        lines.extend(f"- {blocker}" for blocker in report["blockers"])
    lines.append(f"Next action: {report['next_action']}")
    return "\n".join(lines)


def handle_telegram_command(text: str, args: argparse.Namespace) -> str:
    command = compact_spaces(text).split(" ", 1)[0].lower()
    command = command.split("@", 1)[0]
    if command in {"/start", "/help"}:
        return telegram_help_text()
    if command == "/status":
        return build_status_text(Path(args.tracker))
    if command in {"/sync_cal", "/sync-cal"}:
        api_key = os.environ.get("CAL_API_KEY", "").strip()
        if not api_key:
            return "CAL_API_KEY is missing from .env or the environment."
        result = sync_cal_bookings(
            api_key=api_key,
            tracker=Path(args.tracker),
            outbox=Path(args.outbox),
            prospects=known_prospect_index(args),
            statuses=args.status,
            api_version=args.api_version,
        )
        return f"{result['message']}\nUnmatched report: {result['report_path']}"
    if command == "/doctor":
        return build_doctor_text(args)
    return "Unknown command.\n\n" + telegram_help_text()


def cmd_telegram_bot(args: argparse.Namespace) -> int:
    token = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
    allowed_chat_id = os.environ.get("TELEGRAM_CHAT_ID", "").strip()
    if not token or not allowed_chat_id:
        print("TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID are required in .env or the environment.", file=sys.stderr)
        return 2

    print(f"Telegram bot listening for chat {allowed_chat_id}. Press Ctrl+C to stop.")
    offset = int(args.offset or 0)
    while True:
        try:
            updates = fetch_telegram_updates(token, offset=offset, timeout=args.poll_timeout)
            for update in updates:
                offset = max(offset, int(update.get("update_id", 0)) + 1)
                message = update.get("message") or {}
                chat = message.get("chat") or {}
                chat_id = str(chat.get("id", ""))
                text = str(message.get("text") or "")
                if chat_id != allowed_chat_id or not text.startswith("/"):
                    continue
                reply = handle_telegram_command(text, args)
                send_telegram_message(token, chat_id, reply[:4000])
            if args.once:
                break
        except KeyboardInterrupt:
            print("\nTelegram bot stopped.")
            break
        except ApiRequestError as exc:
            print(f"Telegram bot API error: HTTP {exc.status_code}: {compact_spaces(exc.details)}", file=sys.stderr)
            if args.once:
                return 1
            time.sleep(args.error_sleep)
        except Exception as exc:
            print(f"Telegram bot error: {exc}", file=sys.stderr)
            if args.once:
                return 1
            time.sleep(args.error_sleep)
    return 0


def cmd_notify_test(args: argparse.Namespace) -> int:
    message = args.message or "Rocasis sales agent notification test."
    try:
        sent = send_notification(message)
    except ApiRequestError as exc:
        print(f"Telegram notification failed: HTTP {exc.status_code}: {compact_spaces(exc.details)}", file=sys.stderr)
        return 1
    if sent:
        print("Telegram notification sent.")
        return 0
    return 2


def cmd_webhook_cal(args: argparse.Namespace) -> int:
    tracker = Path(args.tracker)
    prospects = known_prospect_index(args)
    secret = os.environ.get("CAL_WEBHOOK_SECRET", "") or args.secret

    class CalWebhookHandler(http.server.BaseHTTPRequestHandler):
        def do_POST(self) -> None:  # noqa: N802
            if secret:
                provided = self.headers.get("x-rocasis-webhook-secret", "")
                query = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
                provided = provided or (query.get("secret", [""])[0])
                if provided != secret:
                    self.send_response(401)
                    self.end_headers()
                    self.wfile.write(b"unauthorized")
                    return
            length = int(self.headers.get("content-length", "0") or "0")
            raw = self.rfile.read(length)
            try:
                payload = json.loads(raw.decode("utf-8"))
            except json.JSONDecodeError:
                self.send_response(400)
                self.end_headers()
                self.wfile.write(b"invalid json")
                return
            rows = tracker_rows(tracker)
            booking = extract_booking(payload)
            matched = mark_booking_confirmed(rows, prospects, booking, "Confirmed from Cal.com webhook")
            write_tracker(tracker, rows)
            self.send_response(200)
            self.send_header("content-type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"status": "ok", "matched": matched}).encode("utf-8"))

        def log_message(self, fmt: str, *values: object) -> None:
            print(f"{self.address_string()} - {fmt % values}")

    server = http.server.ThreadingHTTPServer((args.host, args.port), CalWebhookHandler)
    print(f"Listening for Cal.com webhooks on http://{args.host}:{args.port}/cal-webhook")
    if secret:
        print("Webhook secret check enabled via x-rocasis-webhook-secret header or ?secret=...")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("Stopping webhook server.")
    finally:
        server.server_close()
    return 0


def cmd_simulate_booking(args: argparse.Namespace) -> int:
    tracker = Path(args.tracker)
    prospects = known_prospect_index(args)
    rows = tracker_rows(tracker)
    booking = {
        "uid": args.booking_uid,
        "start": args.start,
        "attendees": [{"email": args.email, "name": args.name or args.email}],
    }
    before = json.loads(json.dumps(rows))
    matched = mark_booking_confirmed(rows, prospects, booking, "Simulated Cal.com booking")
    if args.commit:
        write_tracker(tracker, rows)
        print(f"Simulated booking committed. Matched confirmations: {matched}")
    else:
        print(f"Dry-run simulated booking. Matched confirmations: {matched}")
        rows = before
    if matched:
        for row in rows:
            if row.get("email", "").lower() == args.email.lower():
                print(f"- {row.get('company')} / {row.get('prospect')} / {row.get('meeting_start')}")
    return 0 if matched else 1


def cmd_status(args: argparse.Namespace) -> int:
    print(build_status_text(Path(args.tracker)))
    return 0


def build_status_text(tracker: Path) -> str:
    rows = tracker_rows(tracker)
    confirmed = [r for r in rows if r.get("status") == "confirmed"]
    sent = [r for r in rows if r.get("status") == "sent"]
    followup_1 = [r for r in rows if r.get("status") == "followup_1"]
    followup_2 = [r for r in rows if r.get("status") == "followup_2"]
    due = due_followups(rows)
    lines = [
        f"Target meetings: {TARGET_MEETINGS}",
        f"Confirmed meetings: {len(confirmed)}",
        f"Sent, awaiting reply/booking: {len(sent)}",
        f"Follow-up 1 active: {len(followup_1)}",
        f"Follow-up 2 active: {len(followup_2)}",
        f"Follow-ups due now: {len(due)}",
        f"Remaining confirmations needed: {max(TARGET_MEETINGS - len(confirmed), 0)}",
    ]
    if confirmed:
        lines.append("Confirmed:")
        for row in confirmed:
            lines.append(f"- {row.get('company')} / {row.get('prospect')} / {row.get('meeting_confirmed_at')}")
    return "\n".join(lines)


def tracker_summary(rows: list[dict[str, str]]) -> dict[str, int]:
    return {
        "confirmed": sum(1 for row in rows if row.get("status") == "confirmed"),
        "sent": sum(1 for row in rows if row.get("status") == "sent"),
        "followup_1": sum(1 for row in rows if row.get("status") == "followup_1"),
        "followup_2": sum(1 for row in rows if row.get("status") == "followup_2"),
        "closed": sum(1 for row in rows if row.get("status") == "closed"),
        "due_followups": len(due_followups(rows)),
    }


def build_doctor_report(args: argparse.Namespace) -> dict:
    html_source = Path(args.html_source)
    xlsx_source = Path(args.xlsx_source)
    tracker = Path(args.tracker)
    outbox = Path(args.outbox)
    rows = tracker_rows(tracker)
    prospects: list[Prospect] = []
    source_error = ""
    try:
        prospects = qualify(load_prospects(html_source, xlsx_source))
    except Exception as exc:
        source_error = str(exc)

    resend_key = os.environ.get("RESEND_API_KEY", "").strip()
    cal_key = os.environ.get("CAL_API_KEY", "").strip()
    summary = tracker_summary(rows)
    resend_live = {}
    if getattr(args, "live", False):
        resend_live = validate_resend_live(resend_key, getattr(args, "sender", ""))
    checks = {
        "html_source_exists": html_source.exists(),
        "xlsx_source_exists": xlsx_source.exists(),
        "source_load_ok": bool(prospects) and not source_error,
        "qualified_prospects": len(prospects),
        "tracker_exists": tracker.exists(),
        "outbox_exists": outbox.exists(),
        "email_previews_exists": (outbox / "email_previews.txt").exists(),
        "selected_prospects_exists": (outbox / "selected_prospects.csv").exists(),
        "resend_key_present": bool(resend_key),
        "resend_key_format_ok": validate_resend_api_key(resend_key) if resend_key else False,
        "resend_live_ok": bool(resend_live.get("ok")) if resend_live else None,
        "resend_live_error": resend_live.get("error", "") if resend_live else "",
        "cal_key_present": bool(cal_key),
        "confirmed_meetings": summary["confirmed"],
        "target_meetings": TARGET_MEETINGS,
        "remaining_meetings": max(TARGET_MEETINGS - summary["confirmed"], 0),
        "sent": summary["sent"],
        "followup_1": summary["followup_1"],
        "followup_2": summary["followup_2"],
        "due_followups": summary["due_followups"],
    }
    blockers = []
    if source_error:
        blockers.append(f"Prospect source error: {source_error}")
    if len(prospects) < TARGET_MEETINGS:
        blockers.append("Not enough qualified prospects for the meeting target.")
    if not checks["resend_key_present"]:
        blockers.append("RESEND_API_KEY is missing.")
    elif not checks["resend_key_format_ok"]:
        blockers.append("RESEND_API_KEY does not look like a Resend key; expected prefix re_.")
    elif resend_live and not resend_live.get("ok"):
        blockers.append(str(resend_live.get("error")))
    if summary["confirmed"] < TARGET_MEETINGS and not summary["sent"]:
        next_action = "Validate Resend, then send the initial outreach wave."
    elif summary["due_followups"]:
        next_action = "Send due follow-ups."
    elif summary["confirmed"] < TARGET_MEETINGS:
        next_action = "Sync Cal.com bookings and continue follow-up cadence."
    else:
        next_action = "Meeting target achieved."
    return {"checks": checks, "blockers": blockers, "next_action": next_action}


def cmd_doctor(args: argparse.Namespace) -> int:
    report = build_doctor_report(args)
    if args.json:
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        checks = report["checks"]
        print("Rocasis Sales Agent Doctor")
        print(f"- Qualified prospects: {checks['qualified_prospects']}")
        print(f"- Confirmed meetings: {checks['confirmed_meetings']}/{checks['target_meetings']}")
        print(f"- Sent awaiting booking: {checks['sent']}")
        print(f"- Follow-ups due: {checks['due_followups']}")
        print(f"- Resend key present: {checks['resend_key_present']}")
        print(f"- Resend key format ok: {checks['resend_key_format_ok']}")
        if checks["resend_live_ok"] is not None:
            print(f"- Resend live ok: {checks['resend_live_ok']}")
            if checks["resend_live_error"]:
                print(f"- Resend live error: {checks['resend_live_error']}")
        print(f"- Cal.com key present: {checks['cal_key_present']}")
        if report["blockers"]:
            print("Blockers:")
            for blocker in report["blockers"]:
                print(f"- {blocker}")
        print(f"Next action: {report['next_action']}")
    return 1 if report["blockers"] else 0


def cmd_confirm(args: argparse.Namespace) -> int:
    rows = tracker_rows(Path(args.tracker))
    index = {row.get("email", "").lower(): row for row in rows}
    key = args.email.lower()
    row = index.get(key)
    if not row:
        row = {"email": key}
        rows.append(row)
    row.update(
        {
            "status": "confirmed",
            "booking_uid": args.booking_uid or row.get("booking_uid", ""),
            "meeting_start": args.meeting_start or row.get("meeting_start", ""),
            "meeting_confirmed_at": now_iso(),
            "notes": args.notes or row.get("notes", ""),
        }
    )
    write_tracker(Path(args.tracker), rows)
    print(f"Marked confirmed: {args.email}")
    return 0


def add_common(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--html-source", default=str(DEFAULT_HTML))
    parser.add_argument("--xlsx-source", default=str(DEFAULT_XLSX))
    parser.add_argument("--tracker", default=str(DEFAULT_TRACKER))
    parser.add_argument("--outbox", default=str(DEFAULT_OUTBOX))
    parser.add_argument("--limit", type=int, default=20)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Rocasis sales agent for Cal.com meetings.")
    sub = parser.add_subparsers(dest="cmd", required=True)

    prepare = sub.add_parser("prepare", help="Build shortlist and email previews.")
    add_common(prepare)
    prepare.set_defaults(func=cmd_prepare)

    send = sub.add_parser("send", help="Send selected outreach through Resend.")
    add_common(send)
    send.add_argument("--from", dest="sender", default="")
    send.add_argument("--reply-to", default=DEFAULT_REPLY_TO)
    send.add_argument("--delay-seconds", type=float, default=1.0)
    send.add_argument("--send", action="store_true", help="Actually send. Omitted means dry-run.")
    send.set_defaults(func=cmd_send)

    followups = sub.add_parser("followups", help="Generate or send due follow-ups.")
    followups.add_argument("--tracker", default=str(DEFAULT_TRACKER))
    followups.add_argument("--outbox", default=str(DEFAULT_OUTBOX))
    followups.add_argument("--limit", type=int, default=0)
    followups.add_argument("--from", dest="sender", default="")
    followups.add_argument("--reply-to", default=DEFAULT_REPLY_TO)
    followups.add_argument("--delay-seconds", type=float, default=1.0)
    followups.add_argument("--send", action="store_true", help="Actually send. Omitted means dry-run.")
    followups.set_defaults(func=cmd_followups)

    check_resend = sub.add_parser("check-resend", help="List Resend domains and verify optional sender domain.")
    check_resend.add_argument("--from", dest="sender", default="")
    check_resend.set_defaults(func=cmd_check_resend)

    sync_resend = sub.add_parser("sync-resend", help="Sync sent email statuses from Resend.")
    sync_resend.add_argument("--tracker", default=str(DEFAULT_TRACKER))
    sync_resend.add_argument("--delay-seconds", type=float, default=0.2)
    sync_resend.add_argument("--quiet", action="store_true")
    sync_resend.set_defaults(func=cmd_sync_resend)

    sync_cal = sub.add_parser("sync-cal", help="Sync matching Cal.com bookings into confirmed tracker state.")
    add_common(sync_cal)
    sync_cal.add_argument("--api-version", default=CAL_API_VERSION)
    sync_cal.add_argument("--status", action="append", default=["upcoming"], help="Cal.com booking status to fetch.")
    sync_cal.add_argument("--notify", action="store_true", help="Send Telegram notification when bookings match.")
    sync_cal.set_defaults(func=cmd_sync_cal)

    notify_test = sub.add_parser("notify-test", help="Send a Telegram test notification.")
    notify_test.add_argument("--message", default="")
    notify_test.set_defaults(func=cmd_notify_test)

    telegram_bot = sub.add_parser("telegram-bot", help="Run an interactive Telegram command bot.")
    add_common(telegram_bot)
    telegram_bot.add_argument("--api-version", default=CAL_API_VERSION)
    telegram_bot.add_argument("--status", action="append", default=["upcoming"], help="Cal.com booking status to fetch.")
    telegram_bot.add_argument("--poll-timeout", type=int, default=25)
    telegram_bot.add_argument("--error-sleep", type=float, default=5.0)
    telegram_bot.add_argument("--offset", type=int, default=0)
    telegram_bot.add_argument("--once", action="store_true", help="Process one poll cycle and exit.")
    telegram_bot.set_defaults(func=cmd_telegram_bot)

    webhook_cal = sub.add_parser("webhook-cal", help="Run a local Cal.com webhook receiver.")
    add_common(webhook_cal)
    webhook_cal.add_argument("--host", default="127.0.0.1")
    webhook_cal.add_argument("--port", type=int, default=8787)
    webhook_cal.add_argument("--secret", default="")
    webhook_cal.set_defaults(func=cmd_webhook_cal)

    simulate_booking = sub.add_parser("simulate-booking", help="Simulate a Cal.com booking against the tracker.")
    add_common(simulate_booking)
    simulate_booking.add_argument("--email", required=True)
    simulate_booking.add_argument("--name", default="")
    simulate_booking.add_argument("--start", default="2026-06-05T15:00:00Z")
    simulate_booking.add_argument("--booking-uid", default="simulated_booking")
    simulate_booking.add_argument("--commit", action="store_true", help="Write the simulated confirmation to tracker.")
    simulate_booking.set_defaults(func=cmd_simulate_booking)

    doctor = sub.add_parser("doctor", help="Audit readiness, blockers, and next action.")
    add_common(doctor)
    doctor.add_argument("--json", action="store_true")
    doctor.add_argument("--live", action="store_true", help="Call external APIs for live credential checks.")
    doctor.add_argument("--from", dest="sender", default="")
    doctor.set_defaults(func=cmd_doctor)

    status = sub.add_parser("status", help="Show meeting target progress.")
    status.add_argument("--tracker", default=str(DEFAULT_TRACKER))
    status.set_defaults(func=cmd_status)

    confirm = sub.add_parser("confirm", help="Mark a prospect as confirmed/booked.")
    confirm.add_argument("--tracker", default=str(DEFAULT_TRACKER))
    confirm.add_argument("--email", required=True)
    confirm.add_argument("--notes", default="")
    confirm.add_argument("--booking-uid", default="")
    confirm.add_argument("--meeting-start", default="")
    confirm.set_defaults(func=cmd_confirm)

    return parser


def main(argv: list[str] | None = None) -> int:
    load_dotenv()
    parser = build_parser()
    args = parser.parse_args(argv)
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
